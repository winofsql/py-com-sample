# python -m pip openpyxl install openpyxl
import openpyxl
import traceback
import sys

try:
    # ****************************
    # 新しいブックを作成
    # ****************************
    Book = openpyxl.Workbook()

    # ****************************
    # 通常一つのシートが作成されています
    # ****************************
    Sheet = Book.worksheets[0]

    # ****************************
    # シート名変更
    # ****************************
    Sheet.title = "Pythonの処理"

    # ****************************
    # セルに値を直接セット
    # ****************************
    for i in range(1, 11):
        Sheet.cell(i, 1, f"処理 : {i}")

    Sheet.merge_cells("C1:F1")
    Sheet.cell(1, 3).value = "結合されたセル"

    Book.save('sample.xlsx')

except Exception:
    traceback.print_exc()
    sys.exit( )

print("処理を終了します")